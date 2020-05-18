from datetime import date, datetime, timedelta
import os
import openpyxl
from django.contrib import messages
from django.core import serializers
from django.http import HttpResponse, response
from django.shortcuts import render, redirect
from django.db.models import Q, QuerySet

from django.template import loader
from openpyxl import Workbook
from django.utils.translation import ugettext_lazy as _
from django.core.mail import send_mail

from django.conf import settings
from fluidsurf.apps.dashboard.models import RegistroCompras, RegistroFotografos, RegistroSurferos
from fluidsurf.apps.home.filters import ProductoFilter, UserFilter, CompraFilter, DenunciaFilter, SolicitudFilter, \
    ZonaFilter
from fluidsurf.apps.home.models import Producto, Compra, Denuncia, WatermarkImage, SolicitudStock, Spot, Continente, \
    Pais, Area
from fluidsurf.apps.users.models import CustomUser
from fluidsurf.apps.helpers.helper import registros_vacios_compras, registros_vacios_fotografos, registros_vacios_surferos

def dashboard(request):
    template = loader.get_template('dashboard/main.html')

    # Si el usuario no tiene permisos de administracion, se le impedira acceder al dashboard.
    if not request.user.is_staff:
        return redirect('/')

    registros_vacios_compras()

    compras_query = RegistroCompras.objects.all().query
    results = QuerySet(query=compras_query, model=RegistroCompras).order_by('-fecha')[:7]
    orderResults = reversed(results)
    json = serializers.serialize('json', orderResults)

    # Bucle que estoy usando para calcular el numero total de usuarios nuevos de esta semana
    purchases_last_week = 0
    for item in results:
        purchases_last_week += int(item.compras)

    registros_vacios_fotografos()

    photo_query = RegistroFotografos.objects.all().query
    results = QuerySet(query=photo_query, model=RegistroFotografos)[:30]
    json_photo = serializers.serialize('json', results)

    # Bucle que estoy usando para calcular el numero total de usuarios nuevos de esta semana
    users_last_week = 0
    for item in results:
        users_last_week += int(item.users)

    registros_vacios_surferos()

    surf_query = RegistroSurferos.objects.all().query
    results = QuerySet(query=surf_query, model=RegistroSurferos)[:30]
    json_surf = serializers.serialize('json', results)

    # Continua el blucle anterior aqui
    for item in results:
        users_last_week += int(item.users)

    productos = Producto.objects.filter().all()
    europa = Producto.objects.filter(spot__nombre='Europa').all()
    oceania = Producto.objects.filter(spot__nombre='Oceania').all()
    africa = Producto.objects.filter(spot__nombre='Africa').all()
    asia = Producto.objects.filter(spot__nombre='Asia').all()
    america = Producto.objects.filter(Q(spot__nombre='America del Norte') | Q(spot__nombre='America del Sur')).all()
    compras = Compra.objects.filter().all()
    usuarios = CustomUser.objects.filter().all()

    fotografos = CustomUser.objects.filter(tipo_de_usuario='FOTOGRAFO').all()
    surferos = CustomUser.objects.filter(tipo_de_usuario='SURFERO').all()

    context = {
        'productos': productos,
        'europa': europa,
        'oceania': oceania,
        'africa': africa,
        'asia': asia,
        'america': america,
        'compras': compras,
        'usuarios': usuarios,
        'fotografos': fotografos,
        'surferos': surferos,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count(),
        'array_compras': json,
        'array_photo': json_photo,
        'array_surf': json_surf,
        'purchases_last_week': purchases_last_week,
        'users_last_week': users_last_week
    }

    return HttpResponse(template.render(context, request))


def productos(request):
    template = loader.get_template('dashboard/productos.html')

    productos = Producto.objects.filter().all()

    prod_filter = ProductoFilter(request.GET, queryset=productos)

    if request.method == "POST":
        if 'import' in request.POST:
            excel_file = request.FILES["fileInput"]
            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb['Sheet']

            # Estas dos lineas son para ignorar la primera fila del excel, ya que contiene los
            # nombres de las columnas
            elements = worksheet.iter_rows()
            next(elements)

            for row in elements:
                producto = Producto()
                producto.id = str(row[0].value)
                producto.nombre = str(row[1].value)
                producto.precio = str(row[2].value)
                producto.fecha = str(row[3].value.strftime("%Y-%m-%d"))
                producto.spot = str(row[4].value)
                producto.stock = str(row[5].value)

                producto.imagen0 = str(row[7].value)[7:]
                producto.imagen1 = str(row[8].value)[7:]
                producto.imagen2 = str(row[9].value)[7:]
                producto.imagen3 = str(row[10].value)[7:]
                producto.imagen4 = str(row[11].value)[7:]
                producto.imagen5 = str(row[12].value)[7:]
                producto.imagen6 = str(row[13].value)[7:]
                producto.imagen7 = str(row[14].value)[7:]
                producto.imagen8 = str(row[15].value)[7:]
                producto.imagen9 = str(row[16].value)[7:]

                usuario = CustomUser.objects.filter(username=row[6].value).first()
                producto.user = usuario

                producto.save()
            messages.success(request, 'Productos importados correctamente!')
        else:
            workbook = Workbook()
            sheet = workbook.active

            sheet.append(["ID", "Nombre", "Precio", "Fecha", "Spot", "Stock", "Usuario", 'imagen0', 'imagen1', 'imagen2'
                             , 'imagen3', 'imagen4', 'imagen5', 'imagen6', 'imagen7', 'imagen8', 'imagen9'])

            for p in productos:
                data = [p.id, p.nombre, p.precio, p.fecha, p.spot, p.stock, p.user.username]
                for i in range(10):
                    if getattr(p, 'imagen' + str(i)):
                        data.append(getattr(p, 'imagen' + str(i)).url)
                    else:
                        break
                sheet.append(data)

            if not os.path.isdir("spreadsheets"):
                os.makedirs("spreadsheets")

            workbook.save(filename="spreadsheets/productos" + str(date.today()) + ".xlsx")
            file_path = os.path.join("spreadsheets/productos" + str(date.today()) + ".xlsx")
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                    return response
    context = {
        'filter': prod_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def compras(request):
    template = loader.get_template('dashboard/compras.html')

    compras = Compra.objects.filter().all()

    compras_filter = CompraFilter(request.GET, queryset=compras)

    if request.method == "POST":
        if 'import' in request.POST:
            excel_file = request.FILES["fileInput"]
            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb['Sheet']

            # Estas dos lineas son para ignorar la primera fila del excel, ya que contiene los
            # nombres de las columnas
            elements = worksheet.iter_rows()
            next(elements)

            for row in elements:
                compra = Compra()
                compra.id = str(row[0].value)

                comprador = CustomUser.objects.filter(username=row[1].value).first()
                compra.comprador = comprador

                vendedor = CustomUser.objects.filter(username=row[2].value).first()
                compra.vendedor = vendedor

                producto = Producto.objects.filter(id=row[3].value).first()
                compra.producto = producto

                compra.fecha = str(row[4].value.strftime("%Y-%m-%d"))

                compra.save()
            messages.success(request, 'Compras importadas correctamente!')
        else:
            workbook = Workbook()
            sheet = workbook.active

            sheet.append(["ID", "Comprador", "Vendedor", "Producto", "Fecha"])

            for c in compras:
                data = [c.id, c.comprador.username, c.vendedor.username, c.producto.id, c.fecha]
                sheet.append(data)

            if not os.path.isdir("spreadsheets"):
                os.makedirs("spreadsheets")

            workbook.save(filename="spreadsheets/compras" + str(date.today()) + ".xlsx")
            file_path = os.path.join("spreadsheets/compras" + str(date.today()) + ".xlsx")
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                    return response

    context = {
        'filter': compras_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def zonas(request):
    template = loader.get_template('dashboard/zonas.html')

    zonas = Spot.objects.filter().all()
    zonas_filter = ZonaFilter(request.GET, queryset=zonas)

    if request.method == "POST":
        if 'import' in request.POST:
            excel_file = request.FILES["fileInput"]
            wb = openpyxl.load_workbook(excel_file)

            if wb['DEFINITIVO']:
                worksheet = wb['DEFINITIVO']
            else:
                worksheet = wb['Sheet']

            # Estas dos lineas son para ignorar la primera fila del excel, ya que contiene los
            # nombres de las columnas
            elements = worksheet.iter_rows()
            next(elements)

            for row in elements:

                if not row[0].value is None:
                    xls_cont = row[1].value
                    xls_pais = row[3].value
                    xls_area = row[4].value
                    xls_spot = row[5].value
                    # Comprobante de continentes
                    qs = Continente.objects.filter(nombre=xls_cont)
                    if not qs:
                        continente = Continente()
                        continente.nombre = xls_cont
                        continente.save()
                    # Comprobante de paises
                    qs = Pais.objects.filter(nombre=xls_pais, continente__nombre=xls_cont)
                    if not qs:
                        cont = Continente.objects.filter(nombre=xls_cont).first()
                        pais = Pais()
                        pais.nombre = xls_pais
                        pais.continente = cont
                        pais.save()
                    # Comprobante de areas
                    qs = Area.objects.filter(nombre=xls_area, pais__nombre=xls_pais)
                    if not qs:
                        pais = Pais.objects.filter(nombre=xls_pais).first()
                        area = Area()
                        if xls_area is None: xls_area = ''
                        area.nombre = xls_area
                        area.pais = pais
                        area.save()
                    # Comprobante de spots
                    qs = Spot.objects.filter(nombre=xls_spot, area__nombre=xls_area)
                    if not qs:
                        area = Area.objects.filter(nombre=xls_area).first()
                        spot = Spot()
                        spot.nombre = xls_spot
                        spot.area = area
                        spot.save()
    context = {
        'filter': zonas_filter
    }

    return HttpResponse(template.render(context, request))


def usuarios(request):
    template = loader.get_template('dashboard/usuarios.html')

    usuarios = CustomUser.objects.filter().all()

    user_filter = UserFilter(request.GET, queryset=usuarios)

    if request.method == "POST":
        if 'import' in request.POST:
            excel_file = request.FILES["fileInput"]
            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb['Sheet']

            # Estas dos lineas son para ignorar la primera fila del excel, ya que contiene los
            # nombres de las columnas
            elements = worksheet.iter_rows()
            next(elements)

            for row in elements:
                usuario = CustomUser()
                usuario.id = str(row[0].value)
                usuario.username = str(row[1].value)
                usuario.password = str(row[2].value)

                if row[3].value is not None:
                    usuario.first_name = str(row[3].value)
                else:
                    usuario.first_name = ""
                if row[4].value is not None:
                    usuario.last_name = str(row[4].value)
                else:
                    usuario.last_name = ""
                if row[5].value is not None:
                    usuario.email = str(row[5].value)
                else:
                    usuario.email = ""

                usuario.is_active = str(row[6].value)
                usuario.is_staff = str(row[7].value)
                usuario.is_superuser = str(row[8].value)
                usuario.tipo_de_usuario = str(row[9].value)

                usuario.profile_pic = str(row[10].value)[7:]
                usuario.main_pic = str(row[11].value)[7:]

                if row[12].value is not None:
                    usuario.zona = str(row[12].value)
                else:
                    usuario.zona = ""

                if row[13].value is not None:
                    usuario.pais = str(row[13].value)
                else:
                    usuario.pais = ""

                if row[14].value is not None:
                    usuario.alias = str(row[14].value)
                else:
                    usuario.alias = ""

                if row[15].value is not None:
                    usuario.CV = str(row[15].value)
                else:
                    usuario.cv = ""

                usuario.save()
            messages.success(request, 'Usuarios importados correctamente!')

        else:
            workbook = Workbook()
            sheet = workbook.active

            sheet.append(["ID", "Username", "Contraseña", "Nombre", "Apellidos", "Email", "Activo", "Staff", 'Admin',
                          'Tipo De Usuario', 'Foto de perfil', 'Foto destacada', 'Zona', 'Pais', 'Alias de Fotografo',
                          'CV de Fotografo'])

            for u in usuarios:
                if u.first_name is not None:
                    nombre = u.first_name
                else:
                    nombre = ""
                if u.last_name is not None:
                    apellidos = u.last_name
                else:
                    apellidos = ""
                if u.email is not None:
                    email = u.email
                else:
                    email = ""
                if u.zona is not None:
                    zona = u.zona
                else:
                    zona = ""
                if u.pais is not None:
                    pais = u.pais
                else:
                    pais = ""
                if u.profile_pic:
                    ppic = u.profile_pic.url
                else:
                    ppic = ""
                if u.main_pic:
                    mpic = u.main_pic.url
                else:
                    mpic = ""
                if u.alias:
                    alias = u.alias
                else:
                    alias = ""
                if u.CV:
                    CV = u.CV
                else:
                    CVs = ""

                data = [u.id, u.username, u.password, nombre, apellidos, email, u.is_active, u.is_staff, u.is_superuser,
                        u.tipo_de_usuario, ppic, mpic, zona, pais, alias, CV]
                sheet.append(data)

            if not os.path.isdir("spreadsheets"):
                os.makedirs("spreadsheets")

            workbook.save(filename="spreadsheets/usuarios" + str(date.today()) + ".xlsx")
            file_path = os.path.join("spreadsheets/usuarios" + str(date.today()) + ".xlsx")
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                    return response

    context = {
        'filter': user_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def perfil(request, id=''):
    template = loader.get_template('dashboard/perfil.html')

    usuario = CustomUser.objects.filter(id=id).first()
    wishlist = len(usuario.wishlist.split(',')) - 1

    productos = Producto.objects.filter(user=usuario).all()
    compras = Compra.objects.filter(comprador=usuario).all()
    ventas = Compra.objects.filter(vendedor=usuario).all()
    denuncias = Denuncia.objects.filter(receptor=usuario).all()

    if request.method == "POST":
        # Togglea si el usuario esta activo o no
        usuario.is_active = not usuario.is_active

        usuario.save()

    context = {
        'usuario': usuario,
        'wishlist': wishlist,
        'productos': productos,
        'compras': compras,
        'ventas': ventas,
        'denuncias': denuncias,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def denuncias(request):
    template = loader.get_template('dashboard/denuncias.html')

    denuncias = Denuncia.objects.filter().all()
    denuncias_filter = DenunciaFilter(request.GET, queryset=denuncias)

    print(denuncias_filter)

    if request.method == "POST":
        denuncia = Denuncia.objects.filter(id=request.POST.get('borrar')).first()
        denuncia.delete()

    context = {
        'filter': denuncias_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count()
    }

    return HttpResponse(template.render(context, request))


def validar(request):
    template = loader.get_template('dashboard/validar.html')

    usuarios = CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all()
    user_filter = UserFilter(request.GET, queryset=usuarios)

    if request.method == "POST":
        usuario = CustomUser.objects.filter(id=request.POST.get('validar')).first()
        usuario.validado = True
        usuario.save()
        messages.success(request, 'El usuario ' + usuario.username + ' ha sido validado.')

    context = {
        'filter': user_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def solicitud(request):
    template = loader.get_template('dashboard/solicitudes.html')

    solicitudes = SolicitudStock.objects.all()

    solicitud_filter = SolicitudFilter(request.GET, queryset=solicitudes)


    if request.method == "POST":
        solicitud = SolicitudStock.objects.filter(id=request.POST.get('validar')).first()
        solicitudes = SolicitudStock.objects.filter(product=solicitud.product).all()

        to_mail = []
        for solicitud in solicitudes:
            solicitud.product.stock = 1
            solicitud.product.save()

            if solicitud.user.email: to_mail.append(solicitud.user.email) 
            print(to_mail)

            solicitud.delete()
            
        
        subject = _("A product you asked for has been restocked!")
        message = _("The product %s is back in FluidSurf market!" % (solicitud.product))
        from_email = settings.SERVER_EMAIL
        send_mail(subject, message, from_email, [to_mail, settings.SERVER_EMAIL])


        messages.success(request, 'Se ha repuesto el stock correctamente.')
        return redirect('solicitudes')

    context = {
        'filter': solicitud_filter,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))


def watermark(request):
    template = loader.get_template('dashboard/watermark.html')

    watermarks = WatermarkImage.objects.filter().all()

    if request.method == "POST":
        if 'desactivar' in request.POST:
            watermark = WatermarkImage.objects.filter(id=request.POST.get('desactivar')).first()
            watermark.activo = False
            watermark.save()
        elif "activar" in request.POST:
            for watermark in watermarks:
                watermark.activo = False
                watermark.save()

            watermark = WatermarkImage.objects.filter(id=request.POST.get('activar')).first()
            watermark.activo = True
            watermark.save()

            return redirect('watermark')
        else:
            for watermark in watermarks:
                watermark.activo = False
                watermark.save()

            file = request.FILES['docfile']
            watermark = WatermarkImage()
            watermark.activo = True
            watermark.imagen = file
            watermark.save()

            return redirect('watermark')

    context = {
        'watermarks': watermarks,
        'numero': CustomUser.objects.filter(tipo_de_usuario="FOTOGRAFO", validado=False).all().count(),
        'num_solicitudes': SolicitudStock.objects.all().count()
    }

    return HttpResponse(template.render(context, request))
